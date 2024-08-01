import pandas as pd
from flask import Flask, render_template, request, redirect, send_file
import requests, xlsxwriter, datetime
from openpyxl.styles import Border, Side, Font, Alignment
from io import BytesIO
import locale

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('pagamentos.html')

@app.route('/botao', methods=['POST'])
def botao():
    dataAtual = datetime.date.today()
    dataAtualStr = dataAtual.strftime("%Y-%m-%d")

    tipoVenc = request.form['tipoVenc']
    empresa = request.form['Empresa']
    tipoDoc, nomeDoc = "", ""

    if tipoVenc == "0":
        params = {
            "VencAte": dataAtualStr
        }
        tipoDoc = " - Contas vencidas"
    elif tipoVenc == "1":
        params = {
            "VencDe": dataAtualStr,
            "VencAte": dataAtualStr
        }
        tipoDoc = " - Contas a pagar hoje"
    else:
        return "Opção inválida. Por favor, escolha '0' ou '1'."


    
    if empresa == 'Cel Consultoria':
    #cel consultoria:
        Link = 	"https://celconsultoria.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"CelConsultoria4952",
            "IDUsr":	"28",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSUqEQkF+QIU6SUQ+QDpEgkdBOkpCfkE6hYVBfoVGhD5FgUZEgYSNgZGPlpmNl5GViX5ESkVBQUk=",
            "Usr":	"89e32a1f-8401-4d72-93a2-ff2af6e05c64"
            }
        nomeDoc = "Cel Consultoria" + tipoDoc

    elif empresa == 'Club Gas':
    #club gas
        Link = 	"https://clubgas.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Clubgas6201",
            "IDUsr":	"60",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBRkdKgj5HSoA6RYJEhDpEQUaCOoB+Qj46Pj6CgoGASn5CRkpKgY2ZgIZ+lkZBPkBGPg==",
            "Usr":	"679d079b-5d4e-426d-ba30-00ddcb9a3699"
            }
        nomeDoc = "Club Gas" + tipoDoc

    elif empresa == 'Elevaton':
    #Elevaton
        Link = 	"https://elevaton.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Elevation7111",
            "IDUsr":	"80",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSoJKhD5HQYQ6hYCBQTpESkR+OoCCgkk6hYVCRkJERklKR4JEhI2Emn6XiZGPR0BAQEk+",
            "Usr":	"9d9e072e-fbc2-494a-bdd8-ff36346897d4"
            }
        nomeDoc = "Elevaton" + tipoDoc

    elif empresa == 'Arx':
    #grupoarx
        Link = 	"https://grupoarx.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"GrupoARX3856",
            "IDUsr":	"30",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSUZKSkdBRkE6fj4+hDpERIE+On6ESUY6gkRBSkmEgIRBgUCChpWZkpF+lZxCSUVGQj4=",
            "Usr":	"86997262-a00e-44c0-ae86-d4298ebe2c1d"
            }
        nomeDoc = "Arx" + tipoDoc

    elif empresa == 'Instituo Afeto':
    #instituto afeto
        Link = 	"https://institutoafeto.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"InstitutoAfeto5696",
            "IDUsr":	"22",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBRoFBgYBGSYE6gklHgTpEQEpAOn5KPkI6RYJAhIKCR0BBgUlJiY+Wl4mXmZeRfoWEl5FFRkpGQUE=",
            "Usr":	"6c2cb68c-d87c-4191-a903-5d1edd712c88"
            }
        nomeDoc = "Instituto afeto" + tipoDoc

    elif empresa == 'Otica Vida':
    #óticas vida
        Link = 	"https://oticasvida.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"OticasVida7802",
            "IDUsr":	"20",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBR0RKRYFKQEc6gEE+gjpEhUeFOklARkI6hEpHRYJARklJgoJFkZeJgX6WmomCfkdJPkFBPg==",
            "Usr":	"7495c917-b20d-4f7f-8163-e975d1688dd5"
            }
        nomeDoc = "Otica Vida" + tipoDoc

    elif empresa == 'Porto Leal':
    #porto leal 
        Link = 	"https://portoleal.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"PortoLeal5457",
            "IDUsr":	"26",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSYCBhH5Kfko6gIJKRzpEQUJJOkqCPkQ6R35BRElBhH4+SoWBkpGVl5GNhH6NRURFR0FG",
            "Usr":	"8bcea9a9-bd97-4238-9d04-7a2482ea09fc"
            }
        nomeDoc = "Porto Leal" + tipoDoc

    elif empresa == 'Protege Car':
    #protege car 
        Link = 	"https://protegecar.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",	
            "CN":		"ProjetoCar2680",
            "IDUsr":	"25",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSUI+RIGEhYI6gIJJRDpEhYGBOklERoI6RIJEgoGBREqESkmCkpWRioSXkYF+lUFGST5BRQ==",
            "Usr":	"8304cefd-bd84-4fcc-846d-4d4dcc49e98d"
            }
        nomeDoc = "Protege Car" + tipoDoc

    elif empresa == 'Projeto Verde':
    #projeto verde
        Link = 	"https://projetoverde.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"ProjetoVerde2574"	,
            "IDUsr":	"18",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBfoSCgEKERX46hEeBPjpEgX5EOoBEQkQ6gkJCSkJGR0GCR4CCkpWRioSXkZqElYKEQUVHREBJ",
            "Usr":	"aedb3e5a-e7c0-4ca4-b434-d3393672d7bd"
            }
        nomeDoc = "Projeto Verde" + tipoDoc

    elif empresa == 'Raio':
    #raio
        Link = 	"https://raio.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Raio2899",
            "IDUsr":	"22",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBhIBJfj6AgkI6R0V+QTpERkJAOklGQkQ6RUeCgj6CR0qESUZClX6JkUFJSkpBQQ==",
            "Usr":	"eb8a0bd3-75a2-4631-8634-57dd0d79e863"
            }
        nomeDoc = "Raio" + tipoDoc

    elif empresa == 'Speed':
    #speed
        Link = 	"https://speed.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Speed6427",
            "IDUsr":	"80",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBfkdBQUVCRD46PoFASjpEgUJBOkmBSUU6QURASj5EQkREhUCFlpKEhIJGREFHST4=",
            "Usr":	"a7225340-0c19-4c32-8c85-241904344f1f"
            }
        nomeDoc = "Speed" + tipoDoc

    elif empresa == 'Stone':
    #stone
        Link = 	"https://stone.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Stone4002",
            "IDUsr":	"61",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBRYFJQoWAQoA6PkKBQDpERH5COkqFRoA6PkSERUFGSn5GREmElpeRj4REPj5BRkA=",
            "Usr":	"5c83fb3b-03c1-44a3-9f6b-04e5269a648e"
            }
        nomeDoc = "Stone" + tipoDoc

    elif empresa == 'Valle':
    #valle
        Link = 	"https://valle.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"Valle4155",
            "IDUsr":	"22",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBRoGCgYFJPkA6R35GgTpEQklKOoBJgYA6SoFFQkdHSUVJhYJHmn6NjYREQEVFQUE=",
            "Usr":	"6cdcc801-7a6c-4389-b8cb-9c5377858fd7"
            }
        nomeDoc = "Valle" + tipoDoc

    elif empresa == 'Protege Todos':
    #protege todos
        Link = 	"https://protegetodos.kamino.tech/api/financeiro/pagamento/lista/paginada"	
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":		"ProtegeTodos4726",
            "IDUsr":	"14",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBhUpEREZCR0k6gEJFgjpESUlAOn5FSkk6hURFPkKASUCFRX5FkpWRl4SGhJeRgpGWREdBRkBE",
            "Usr":	"f9446378-b35d-4881-a598-f4503b81f5a5"
            }
        nomeDoc = "Protege Todos" + tipoDoc

    elif empresa == 'Unymos':
    #Unymos
        Link = "https://unymos.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"UNYMOSGESTAOESI5245",
            "IDUsr":	"72",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBQX5CgYRAgkc6RkRHhDpEgUA+Oko+hUo6REKAgICAQoRCgX6CmY+ejpGWhoSWl36RhJaJRUFERUdB",
            "Usr":	"2a3ce1d7-647e-4c10-90f9-43bbbb3e3cad"
            }
        nomeDoc = "Unymos" + tipoDoc

    elif empresa == 'Auto Nacional':
    #Auto Nacional
        Link = "https://autonacional.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"AUTONACIONALASS3118",
            "IDUsr":	"62",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSoKEPkKCgkc6RUc+RDpEhYSFOkpKgoU6SX6EPoSBR0FJhEpAfpmXkY9+gYmRj36NfpaWQkBASUZB",
            "Usr":	"9de03dd7-5704-4fef-99df-8ae0ec728e91"
            }
        nomeDoc = "Auto Nacional" + tipoDoc

    elif empresa == 'wr rastreamento':
    #wr rastreamento
        Link = "https://wrrastreamento.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"WRRASTREAMENTOE5125",
            "IDUsr":	"8",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBgoKCRkdKhYA6gEp+RDpEfoBKOklJSYU6hD5HhIFCgoKCfoCEm5WVfpaXlYR+joSPl5GERUBBRUk=",
            "Usr":	"ddd679fb-b9a4-4ab9-888f-e07ec3dddabe"
            }
        nomeDoc = "Wr Rastreamento" + tipoDoc

    elif empresa == 'New Car':
    #newcar
        Link = "https://newcar.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"NEWCARPV7762",
            "IDUsr":	"59",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBgD5+R36EgYI6fkeERTpEQUF+OoBBSX46QICFQIFCREGFfkU+j4SbgX6VkppHR0ZBRUo=",
            "Usr":	"b0a7aecd-a7e5-422a-b28a-1bf1c342fa50"
            }
        nomeDoc = "New Car" + tipoDoc

    elif empresa == 'Altis':
    #Altis
        Link = "https://altis.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"ALTIS2447",
            "IDUsr":	"11",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBSkY+PoFERkY6SkqEgTpEgYRBOkpAQEk6QkBBhIJARn6CRkp+fo2XiZZBRERHQEA=",
            "Usr":	"9600c466-99ec-4ce2-9118-312ed16ad69a"
            }
        nomeDoc = "Altis" + tipoDoc

    elif empresa == 'Protege Car Associação':
    #protege car Associacoes
        Link = "https://protegecarassociacao.kamino.tech/api/financeiro/pagamento/lista/paginada"
        headers = {
            "App":	"b929d3c0-eaac-409f-ab8b-10b86e2c5cec",
            "CN":	"PROTEGECARASSOC2242",
            "IDUsr":	"13",
            "Hash":	"gEpBSoJCgT46hH5+gTpEPkqFOn6ASYA6QD6ASUaEQYFFgYSBPkJAgIVKPkE6RYRFgTpEQYR+OkpHhUU6hYF+RUFBR0dKRoGBkpWRl4SGhIF+lX6WlpGBQUFEQUBC",
            "Usr":	"031bf902-5e5c-42ea-97f5-fca5227796cc"
            }
        nomeDoc = "Protege car" + tipoDoc

    TITULO = ("ARX", "Alguma coisa aqui", nomeDoc)

    try:
        response = requests.get(Link, params=params, headers=headers)
        response.raise_for_status()   
    except requests.exceptions.RequestException as e:
        return f"Erro ao fazer a requisição da API: {e}"

    data = response.json()


    keys_to_extract = [
        "Pessoa.NomeExibicao",
        "ContaClassificacao.Nome",
        "NroNotaFiscal",
        "DescricaoFormaPagamento",
        "Obs",
        "ValorVencimento",
        "_DataVencimento"
    ]

    extracted_data = []


    for dados in data['Dados']:
        extracted_item = {}
        for key in keys_to_extract:
            current_level = dados
            nested_keys = key.split('.')
            for nested_key in nested_keys:
                if current_level is None or nested_key not in current_level:
                    current_level = None
                    break
                current_level = current_level.get(nested_key)
            if current_level is not None:
                extracted_item[key] = current_level
        extracted_data.append(extracted_item)


    df = pd.DataFrame(extracted_data, columns=keys_to_extract)

    df.rename(columns={
        'Pessoa.NomeExibicao': 'Nome Pessoa',
        'ContaClassificacao.Nome': 'Conta',
        'NroNotaFiscal': 'Nota Fiscal',
        'DescricaoFormaPagamento': 'Forma De Pagamento',
        'Obs': 'Observação',
        'ValorVencimento': 'Valor',
        '_DataVencimento': 'Data de Vencimento'
    }, inplace=True)


    df['Data de Vencimento'] = pd.to_datetime(df['Data de Vencimento'], errors='coerce')

    df['Data de Vencimento'] = df['Data de Vencimento'].dt.strftime('%d/%m/%Y')


    excel_data = BytesIO()

    with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Planilha 1', index=False)

        wb = writer.book
        ws = wb['Planilha 1']

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Obtendo o objeto workbook e worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Formato de moeda para a coluna 'Valor'
        currency_format = workbook.add_format({'num_format': '$#,##0.00'})
        worksheet.set_column('C:C', None, currency_format)

        # Calculando o total
        total = df['Valor'].sum()

        # Escrevendo o total na última linha
        last_row = len(df) + 1
        worksheet.write(last_row, 2, 'Total:', currency_format)
        worksheet.write_formula(last_row, 3, f'=SUM(C2:C{last_row})', currency_format)

        total = 0
        for cell in ws['F']:
            if isinstance(cell.value, (int, float)):
                total += cell.value

        total = locale.currency(total, grouping=True, symbol=False)

        total_cell = ws.cell(row=ws.max_row + 1, column=6)
        total_cell.value = f'R{total}'


        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=7)
            if isinstance(cell.value, (int, float)):
                edit = 'R' + locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
                cell.value = edit


        align = Alignment(horizontal='center', vertical='center')
        ws['A1'].alignment = align

        
        ws.insert_rows(1)
        ws['A1'] = "Grupo Arx"

        ws.insert_rows(1)
        ws['A1'] = "Alguma coisa aqui"

        ws.insert_rows(1)
        ws['A1'] = nomeDoc

        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws.merge_cells('A3:G3')

        ws['A1'].alignment = align
        ws['A2'].alignment = align
        ws['A3'].alignment = align


        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        font = Font(name='Arial', size=12)


        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
                cell.font = font


    wb.save(excel_data)
    excel_data.seek(0)   

    return send_file(excel_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=nomeDoc + ".xlsx")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
